import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import timedelta
import warnings

from plotly.colors import find_intermediate_color, hex_to_rgb

# Page-1 base palette (left→right: red → deep blue)
BASE_BUCKET_COLORS = [
    "#d73027", "#fc8d59", "#fee090", "#c6dbef", "#9ecae1",
    "#6baed6", "#4292c6", "#2171b5", "#084594"
]

def build_20min_bins(max_minutes: float):
    """
    Return (edges, labels_no_prefix, labels_with_prefix) 
    for 20-min bins up to ceil(max/20)*20.
    """
    if pd.isna(max_minutes) or max_minutes <= 0:
        edges = [0, 20]
    else:
        upper = int(np.ceil(max_minutes / 20.0) * 20)
        edges = list(range(0, upper + 20, 20))  # 0,20,...,upper
    
    # Labels like "0-20 min", "20-40 min", ...
    labels_np = [f"{edges[i]}-{edges[i+1]} min" for i in range(len(edges)-1)]
    labels_wp = [f"{i+1}: {labels_np[i]}" for i in range(len(labels_np))]
    
    return edges, labels_np, labels_wp

def lighten_hex(hex_color, factor=0.2):
    """
    Lighten a hex color by blending it with white.
    factor=0 → original color, factor=1 → pure white.
    """
    hex_color = hex_color.lstrip("#")
    r, g, b = [int(hex_color[i:i+2], 16) for i in (0, 2, 4)]
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f"#{r:02x}{g:02x}{b:02x}"


def make_bucket_color_map(labels_with_prefix):
    """
    Always include 9 base labels in the legend.
    If more buckets exist, extend colors by reusing and lightening them.
    """
    base = BASE_BUCKET_COLORS

    # Always force legend to show at least 9 entries
    full_labels = labels_with_prefix.copy()
    while len(full_labels) < 9:
        full_labels.append(f"{len(full_labels)+1}: (unused)")

    n = len(full_labels)
    if n <= len(base):
        colors = base[:n]
    else:
        colors = base.copy()
        for i in range(len(base), n):
            base_color = base[i % len(base)]
            # lighten progressively with each cycle
            factor = 0.2 * ((i // len(base)) + 1)
            factor = min(factor, 0.8)  # cap at 80% lightening
            colors.append(lighten_hex(base_color, factor))

    return {lbl: colors[i] for i, lbl in enumerate(full_labels)}

# Excel export helpers
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from io import BytesIO

# Suppress deprecation warnings during dev (optional but recommended)
warnings.filterwarnings("ignore", category=FutureWarning)

st.set_page_config(layout="wide")

# --- Helper Functions ---
def format_time(minutes):
    """Convert minutes (float) to hh:mm:ss string."""
    seconds = int(minutes * 60)
    return str(timedelta(seconds=seconds))

def calculate_run_rate_excel_like(df):
    df = df.copy()
    
    # --- Handle Date/Time Parsing ---
    if {"YEAR", "MONTH", "DAY", "TIME"}.issubset(df.columns):
        df["SHOT TIME"] = pd.to_datetime(
            df["YEAR"].astype(str) + "-" +
            df["MONTH"].astype(str) + "-" +
            df["DAY"].astype(str) + " " +
            df["TIME"].astype(str),
            errors="coerce"
        )
    elif "SHOT TIME" in df.columns:
        df["SHOT TIME"] = pd.to_datetime(df["SHOT TIME"], errors="coerce")
    else:
        st.error("Input file must contain either 'SHOT TIME' or YEAR/MONTH/DAY/TIME columns.")
        st.stop()

    # --- CT Difference Handling (row-shifted) ---
    df["CT_diff_raw"] = df["SHOT TIME"].diff().dt.total_seconds()
    
    # Shift ACTUAL CT down so each row uses previous cycle’s ACTUAL CT
    df["CT_diff_sec"] = df["ACTUAL CT"].shift()
    
    # Rule: if previous ACTUAL CT == 999.9 → fall back to raw gap
    mask_maxed = df["CT_diff_sec"] == 999.9
    df.loc[mask_maxed, "CT_diff_sec"] = df.loc[mask_maxed, "CT_diff_raw"]
    
    # First row: always use raw gap (or NaN if no diff)
    df.loc[df.index[0], "CT_diff_sec"] = df.loc[df.index[0], "CT_diff_raw"]

    # Mode CT (seconds)
    mode_ct = df["ACTUAL CT"].mode().iloc[0]
    
    # ✅ Use tolerance band (from sidebar or default 10%)
    tolerance = st.session_state.get("tolerance", 0.10)   # default 10% if not set
    lower_limit = mode_ct * (1 - tolerance)
    upper_limit = mode_ct * (1 + tolerance)

    # STOP flag (all potential stops)
    df["STOP_FLAG"] = np.where(
        (df["CT_diff_sec"].notna()) &
        ((df["CT_diff_sec"] < lower_limit) | (df["CT_diff_sec"] > upper_limit)) &
        (df["CT_diff_sec"] <= 28800),  # ignore > 8 hours gaps
        1, 0
    )
    df.loc[df.index[0], "STOP_FLAG"] = 0

    # Back-to-back stop adjustment (for stop count)
    df["STOP_ADJ"] = df["STOP_FLAG"]
    df.loc[(df["STOP_FLAG"] == 1) & (df["STOP_FLAG"].shift(fill_value=0) == 1), "STOP_ADJ"] = 0

    # Events (first in sequence = true stop event)
    df["STOP_EVENT"] = (df["STOP_ADJ"].shift(fill_value=0) == 0) & (df["STOP_ADJ"] == 1)

    # --- Core Metrics ---
    total_shots = len(df)
    
    # Normal shots = CT_diff_sec inside band
    normal_shots = ((df["CT_diff_sec"] >= lower_limit) & (df["CT_diff_sec"] <= upper_limit)).sum()
    
    # Bad shots = outside band (but still counted as shots)
    bad_shots = total_shots - normal_shots
    
    # Stop events = first flagged stop in a sequence
    stop_events = df["STOP_EVENT"].sum()

    # --- Time-based Calculations ---
    total_runtime = (df["SHOT TIME"].max() - df["SHOT TIME"].min()).total_seconds() / 60  # minutes
    run_hours = total_runtime / 60

    # Downtime = sum of ALL stop intervals (even back-to-back)
    downtime = df.loc[df["STOP_FLAG"] == 1, "CT_diff_sec"].sum() / 60  # minutes

    # Production time = runtime - downtime
    production_time = total_runtime - downtime

    gross_rate = total_shots / run_hours if run_hours else None
    net_rate = normal_shots / run_hours if run_hours else None
    efficiency = normal_shots / total_shots if total_shots else None

    # Each run = time between two stop events (using STOP_ADJ to collapse back-to-back)
    # --- Continuous Run Durations ---
    df["RUN_GROUP"] = df["STOP_ADJ"].cumsum()
    
    run_durations = (
        df.groupby("RUN_GROUP")
          .apply(lambda g: g["CT_diff_sec"].sum() / 60)  # minutes
          .reset_index(name="RUN_DURATION")
    )
    
    # ✅ Add run end timestamp for each group right away
    run_end_times = df.groupby("RUN_GROUP")["SHOT TIME"].max().reset_index(name="RUN_END")
    run_durations = run_durations.merge(run_end_times, on="RUN_GROUP", how="left")
    
    # Remove first run if it starts with a stop (edge case)
    run_durations = run_durations[run_durations["RUN_DURATION"] > 0]
    
    # 🔹 Remove last run if it didn’t end in a stop (incomplete run)
    if df["STOP_ADJ"].iloc[-1] == 0:
        last_group = df["RUN_GROUP"].iloc[-1]
        run_durations = run_durations[run_durations["RUN_GROUP"] != last_group]

    # --- Assign buckets (dynamic 20-min bins) ---
    max_minutes = run_durations["RUN_DURATION"].max()
    
    # ✅ Safety clamp – cap at 240 min (4 hrs) so buckets don’t explode
    if pd.notna(max_minutes):
        max_minutes = min(max_minutes, 240)
    
    edges, labels_np, labels_wp = build_20min_bins(max_minutes)
    
    run_durations["TIME_BUCKET_RAW"] = pd.cut(
        run_durations["RUN_DURATION"],
        bins=edges,
        right=False,   # left-inclusive, right-exclusive
        labels=labels_np
    )
    
    map_np_to_wp = {np_lbl: wp_lbl for np_lbl, wp_lbl in zip(labels_np, labels_wp)}
    run_durations["TIME_BUCKET"] = run_durations["TIME_BUCKET_RAW"].map(map_np_to_wp)
    
    bucket_order = labels_wp  # keep this for plotting later
    bucket_color_map = make_bucket_color_map(bucket_order)
    
    # 🔹 Drop absurd runs > 8 hours (likely incomplete)
    run_durations = run_durations[run_durations["RUN_DURATION"] <= 480]

    # Bucket counts for overall distribution
    bucket_counts = run_durations["TIME_BUCKET"].value_counts().sort_index().fillna(0).astype(int)
    bucket_counts.loc["Grand Total"] = bucket_counts.sum()

    # --- Hourly MTTR/MTBF ---
    df["HOUR"] = df["SHOT TIME"].dt.hour
    df["DOWNTIME_MIN"] = np.where(df["STOP_EVENT"], df["CT_diff_sec"]/60, np.nan)
    df["UPTIME_MIN"] = np.where(~df["STOP_EVENT"], df["CT_diff_sec"]/60, np.nan)

    def safe_mtbf(uptime_series, stop_count):
        if stop_count > 0 and uptime_series.notna().any():
            return np.nanmean(uptime_series)
        else:
            return np.nan
    
    hourly = (
        df.groupby("HOUR")
          .apply(lambda g: pd.Series({
              "stops": g["STOP_EVENT"].sum(),
              "mttr": np.nanmean(g["DOWNTIME_MIN"]) if g["DOWNTIME_MIN"].notna().any() else np.nan,
              "mtbf": safe_mtbf(g["UPTIME_MIN"], g["STOP_EVENT"].sum())
          }))
          .reset_index()
    )
    hourly["stability_index"] = (hourly["mtbf"] / (hourly["mtbf"] + hourly["mttr"])) * 100

    return {
        "mode_ct": mode_ct,
        "lower_limit": lower_limit,
        "upper_limit": upper_limit,
        "total_shots": total_shots,
        "normal_shots": normal_shots,
        "bad_shots": bad_shots,
        "stop_events": stop_events,
        "run_hours": run_hours,
        "gross_rate": gross_rate,
        "net_rate": net_rate,
        "efficiency": efficiency,
        "production_time": production_time,
        "downtime": downtime,
        "total_runtime": total_runtime,
        "bucket_counts": bucket_counts,
        "hourly": hourly,
        "df": df,
        "run_durations": run_durations,      # ✅ still there
        "bucket_order": bucket_order,        # ✅ NEW
        "bucket_color_map": bucket_color_map # ✅ NEW
    }

# --- UI ---
st.sidebar.title("Run Rate Report Generator")

uploaded_file = st.sidebar.file_uploader(
    "Upload Run Rate Excel (clean table)", type=["xlsx"]
)

if uploaded_file:
    df = pd.read_excel(uploaded_file)

    # --- Tool selection ---
    if "TOOLING ID" in df.columns:
        selection_column = "TOOLING ID"
    elif "EQUIPMENT CODE" in df.columns:
        selection_column = "EQUIPMENT CODE"
    else:
        st.error("File must contain either 'TOOLING ID' or 'EQUIPMENT CODE'.")
        st.stop()

    tool = st.sidebar.selectbox("Select Tool", df[selection_column].unique())
    date = st.sidebar.date_input(
        "Select Date", pd.to_datetime(df["SHOT TIME"]).dt.date.min()
    )

    page = st.sidebar.radio(
        "Select Page",
        ["📊 Analysis Dashboard", "📂 Raw & Processed Data", "📅 Daily Analysis"]
    )

    # --- Generate Report ---
    if st.sidebar.button("Generate Report"):
        mask = (df[selection_column] == tool)
        df_filtered = df.loc[mask]
    
        if df_filtered.empty:
            st.warning("No data found for this selection.")
        else:
            # Store raw once
            st.session_state["df_raw"] = df_filtered.copy()
    
    # --- Threshold & Tolerance Settings (always active if df_raw exists) ---
    if "df_raw" in st.session_state:
        df_raw = st.session_state["df_raw"].copy()
    
        # 🚨 Stoppage Threshold
        st.sidebar.markdown("### 🚨 Stoppage Threshold Settings")
        mode_ct = st.session_state.get("results", {}).get("mode_ct", df_raw["ACTUAL CT"].mode().iloc[0])
    
        threshold_mode = st.sidebar.radio(
            "Select threshold type:",
            ["Multiple of Mode CT", "Manual (seconds)"],
            horizontal=False,
            key="sidebar_threshold_mode"
        )
    
        if threshold_mode == "Multiple of Mode CT":
            multiplier = st.sidebar.slider(
                "Multiplier of Mode CT",
                min_value=1.0, max_value=5.0, value=2.0, step=0.5,
                key="sidebar_ct_multiplier"
            )
            threshold = mode_ct * multiplier
            threshold_label = f"Mode CT × {multiplier} = {threshold:.2f} sec"
        else:
            default_val = float(mode_ct * 2)
            threshold = st.sidebar.number_input(
                "Manual threshold (seconds)",
                min_value=1.0, value=default_val,
                key="sidebar_manual_threshold"
            )
            threshold_label = f"{threshold:.2f} sec (manual)"
    
        st.session_state["threshold_mode"] = threshold_mode
        st.session_state["threshold"] = threshold
        st.session_state["threshold_label"] = threshold_label
    
        # ⚙️ Cycle Time Tolerance
        st.sidebar.markdown("### ⚙️ Cycle Time Tolerance Settings")
        tolerance = st.sidebar.slider(
            "Tolerance Band (% of Mode CT)",
            min_value=0.01, max_value=0.20, value=0.05, step=0.01,
            help="Defines the ±% around Mode CT to classify normal vs. bad shots"
        )
        st.session_state["tolerance"] = tolerance
    
        # ✅ Recalculate results EVERY rerun (with updated settings)
        st.session_state["results"] = calculate_run_rate_excel_like(df_raw)
    
    # --- Page 1: Analysis Dashboard ---
    if page == "📊 Analysis Dashboard":
        results = st.session_state.get("results", {})
        if not results:
            st.info("👈 Please generate a report first from the sidebar.")
        else:
            # ✅ Daily filter
            df_day = results["df"].copy()
            df_day = df_day[pd.to_datetime(df_day["SHOT TIME"]).dt.date == date]
    
            if df_day.empty:
                st.warning("No data found for this date.")
            else:
                # Re-run the calculations just on df_day
                daily_results = calculate_run_rate_excel_like(df_day)
    
                st.title("📊 Run Rate Report")
                st.subheader(f"Tool: {tool} | Date: {date.strftime('%Y-%m-%d')}")
    
                # --- Shot Counts & Efficiency ---
                st.markdown("### Shot Counts & Efficiency")
                st.table(pd.DataFrame({
                    "Total Shot Count": [daily_results['total_shots']],
                    "Normal Shot Count": [daily_results['normal_shots']],
                    "Bad Shot Count": [daily_results['bad_shots']],
                    "Efficiency": [f"{(daily_results['normal_shots']/daily_results['total_shots'])*100:.2f}%"],
                    "Stop Count": [daily_results['stop_events']]
                }))
    
                # ... and below, use daily_results instead of results
            
            # --- Reliability Metrics ---
            results = st.session_state.get("results", {})
            df_res = results.get("df", pd.DataFrame()).copy()
            stop_events = results.get("stop_events", 0)
            
            if stop_events > 0 and "STOP_EVENT" in df_res.columns:
                df_res = df_res.reset_index(drop=True)  # ensure clean index
            
                # Downtime durations (stop events only)
                downtime_events = df_res.loc[df_res["STOP_EVENT"], "CT_diff_sec"] / 60
                mttr = downtime_events.mean() if not downtime_events.empty else None
            
                # MTBF = total uptime (all CTs) ÷ number of stops
                total_uptime = df_res["CT_diff_sec"].sum() / 60  # minutes
                mtbf = total_uptime / stop_events if stop_events > 0 else None
            
                # Time to First DT
                if df_res["STOP_EVENT"].any():
                    first_stop_idx = df_res.index[df_res["STOP_EVENT"]].min()
                
                    # uptime = sum of CTs until the first stop event
                    uptime_until_first = df_res.loc[:first_stop_idx, "CT_diff_sec"].sum() / 60
                
                    # if there was no valid uptime, still return 0.0 instead of None
                    first_dt = max(uptime_until_first, 0.0)
                else:
                    first_dt = np.nan
            else:
                mttr, mtbf, first_dt = None, None, None
            
            avg_ct = df_res["ACTUAL CT"].mean() if "ACTUAL CT" in df_res.columns else None
            
            reliability_df = pd.DataFrame({
                "Metric": ["MTTR (min)", "MTBF (min)", "Time to First DT (min)", "Avg Cycle Time (sec)"],
                "Value": [
                    f"{mttr:.2f}" if mttr else "N/A",
                    f"{mtbf:.2f}" if mtbf else "N/A",
                    f"{first_dt:.2f}" if first_dt else "N/A",
                    f"{avg_ct:.2f}" if avg_ct else "N/A"
                ]
            })
            
            st.markdown("### Reliability Metrics")
            st.table(reliability_df)
    
            # --- Production & Downtime Summary ---
            st.markdown("### Production & Downtime Summary")
            st.table(pd.DataFrame({
                "Mode CT": [f"{results.get('mode_ct', 0):.2f}"],
                "Lower Limit": [f"{results.get('lower_limit', 0):.2f}"],
                "Upper Limit": [f"{results.get('upper_limit', 0):.2f}"],
                "Production Time (hrs)": [
                    f"{results.get('production_time', 0)/60:.1f} hrs "
                    f"({results.get('production_time', 0)/results.get('total_runtime', 1)*100:.2f}%)"
                ],
                "Downtime (hrs)": [
                    f"{results.get('downtime', 0)/60:.1f} hrs "
                    f"({results.get('downtime', 0)/results.get('total_runtime', 1)*100:.2f}%)"
                ],
                "Total Run Time (hrs)": [f"{results.get('run_hours', 0):.2f}"],
                "Total Stops": [stop_events]
            }))
    
            # --- Visual Analysis ---
            st.subheader("📈 Visual Analysis")
            
            run_durations = results["run_durations"].copy()
            bucket_order = results.get("bucket_order", [])
            bucket_color_map = results.get("bucket_color_map", {})
            
            # 1) Time Bucket Analysis
            bucket_counts = run_durations["TIME_BUCKET"].value_counts().reindex(bucket_order).fillna(0).astype(int)
            total_runs = bucket_counts.sum()
            bucket_df = bucket_counts.reset_index()
            bucket_df.columns = ["Time Bucket", "Occurrences"]
            bucket_df["Percentage"] = (bucket_df["Occurrences"] / total_runs * 100).round(2)
            
            fig_bucket = px.bar(
                bucket_df[bucket_df["Time Bucket"].notna()],
                x="Occurrences", y="Time Bucket",
                orientation="h", text="Occurrences",
                title="Time Bucket Analysis (Continuous Runs Before Stops)",
                category_orders={"Time Bucket": results["bucket_order"]},   # ✅ keep order
                color="Time Bucket",
                color_discrete_map=results["bucket_color_map"],             # ✅ fixed palette
                hover_data={"Occurrences": True, "Percentage": True}
            )
            fig_bucket.update_traces(textposition="outside")
            st.plotly_chart(fig_bucket, use_container_width=True)
            
            with st.expander("📊 Time Bucket Analysis Data Table", expanded=False):
                st.dataframe(bucket_df)
    
            # 2) Time Bucket Trend (group by hour of day instead of week)
    
            if "RUN_END" in run_durations.columns:
                run_durations["HOUR"] = run_durations["RUN_END"].dt.hour
            else:
                run_durations["HOUR"] = -1
            
            trend = run_durations.groupby(["HOUR","TIME_BUCKET"]).size().reset_index(name="count")

            # Ensure all hours + all buckets appear
            hours = list(range(24))
            grid = pd.MultiIndex.from_product([hours, results["bucket_order"]], names=["HOUR","TIME_BUCKET"]).to_frame(index=False)
            trend = grid.merge(trend, on=["HOUR","TIME_BUCKET"], how="left").fillna({"count":0})
            
            fig_tb_trend = px.bar(
                trend, x="HOUR", y="count", color="TIME_BUCKET",
                category_orders={"TIME_BUCKET": results["bucket_order"]},
                color_discrete_map=results["bucket_color_map"],   # ✅ fixed palette
                title="Hourly Time Bucket Trend (Continuous Runs Before Stops)",
                hover_data={"count": True, "HOUR": True}
            )
            fig_tb_trend.update_layout(
                barmode="stack",
                xaxis=dict(title="Hour of Day (0–23)", tickmode="linear", dtick=1, range=[-0.5,23.5]),
                yaxis=dict(title="Number of Runs")
            )
            
            st.plotly_chart(fig_tb_trend, width="stretch")
            
            with st.expander("📊 Hourly Time Bucket Trend Data Table", expanded=False):
                st.dataframe(trend)
    
            # 3) MTTR & MTBF Trend by Hour
            hourly = results.get("hourly", pd.DataFrame()).copy()
            all_hours = pd.DataFrame({"HOUR": list(range(24))})
            hourly = all_hours.merge(hourly, on="HOUR", how="left")
            fig_mt = go.Figure()
            fig_mt.add_trace(go.Scatter(x=hourly["HOUR"], y=hourly["mttr"], mode="lines+markers",
                                        name="MTTR (min)", line=dict(color="red", width=2), yaxis="y"))
            fig_mt.add_trace(go.Scatter(x=hourly["HOUR"], y=hourly["mtbf"], mode="lines+markers",
                                        name="MTBF (min)", line=dict(color="green", width=2, dash="dot"), yaxis="y2"))
            fig_mt.update_layout(title="MTTR & MTBF Trend by Hour",
                                 xaxis=dict(title="Hour of Day (0–23)", tickmode="linear", dtick=1, range=[-0.5,23.5]),
                                 yaxis=dict(title="MTTR (min)", tickfont=dict(color="red"), side="left"),
                                 yaxis2=dict(title="MTBF (min)", tickfont=dict(color="green"), overlaying="y", side="right"),
                                 margin=dict(l=60,r=60,t=60,b=40),
                                 legend=dict(orientation="h", x=0.5, y=-0.25, xanchor="center"))
            st.plotly_chart(fig_mt, width="stretch")
            with st.expander("📊 MTTR & MTBF Data Table", expanded=False):
                st.dataframe(hourly)
    
            # 4) Stability Index
            hourly["stability_index"] = np.where((hourly["stops"] == 0) & (hourly["mtbf"].isna()),
                                                 100, hourly["stability_index"])
            hourly["stability_change_%"] = hourly["stability_index"].pct_change() * 100
            colors = ["gray" if pd.isna(v) else "red" if v <= 50 else "yellow" if v <= 70 else "green" for v in hourly["stability_index"]]
            fig_stability = go.Figure()
            fig_stability.add_trace(go.Scatter(x=hourly["HOUR"], y=hourly["stability_index"],
                                               mode="lines+markers", name="Stability Index (%)",
                                               line=dict(color="blue", width=2), marker=dict(color=colors, size=8)))
            for y0,y1,c in [(0,50,"red"),(50,70,"yellow"),(70,100,"green")]:
                fig_stability.add_shape(type="rect", x0=-0.5, x1=23.5, y0=y0, y1=y1,
                                        fillcolor=c, opacity=0.1, line_width=0, yref="y")
            fig_stability.update_layout(title="Stability Index by Hour",
                                        xaxis=dict(title="Hour of Day (0–23)", tickmode="linear", dtick=1, range=[-0.5,23.5]),
                                        yaxis=dict(title="Stability Index (%)", range=[0,100], side="left"),
                                        margin=dict(l=60,r=60,t=60,b=40),
                                        legend=dict(orientation="h", x=0.5, y=-0.25, xanchor="center"))
            st.plotly_chart(fig_stability, width="stretch")
            with st.expander("📊 Stability Index Data Table", expanded=False):
                table_data = hourly[["HOUR","stability_index","stability_change_%","mttr","mtbf","stops"]].copy()
                table_data.rename(columns={
                    "HOUR":"Hour",
                    "stability_index":"Stability Index (%)",
                    "stability_change_%":"Change vs Prev Hour (%)",
                    "mttr":"MTTR (min)",
                    "mtbf":"MTBF (min)",
                    "stops":"Stop Count"
                }, inplace=True)
            
                # Highlight only the Stability Index column
                def highlight_stability(val):
                    if pd.isna(val):
                        return ""
                    elif val <= 50:
                        return "background-color: rgba(255, 0, 0, 0.3);"   # soft red
                    elif val <= 70:
                        return "background-color: rgba(255, 255, 0, 0.3);" # soft yellow
                    else:
                        return ""
                
                st.dataframe(
                    table_data.style
                    .applymap(highlight_stability, subset=["Stability Index (%)"])
                    .format({
                        "Stability Index (%)": "{:.2f}",
                        "Change vs Prev Hour (%)": "{:+.2f}%",
                        "MTTR (min)": "{:.2f}",
                        "MTBF (min)": "{:.2f}"
                    })
                )
    
            st.markdown("""
            **ℹ️ Stability Index Formula**
            - Stability Index (%) = (MTBF / (MTBF + MTTR)) × 100
            - If no stoppages occur in an hour, Stability Index is forced to **100%**
            - Alert Zones:
              - 🟥 0–50% → High Risk (Frequent stoppages with long recovery times. Production is highly unstable.)
              - 🟨 50–70% → Medium Risk (Minor but frequent stoppages or slower-than-normal recoveries. Production flow is inconsistent and requires attention to prevent escalation.)
              - 🟩 70–100% → Low Risk (stable operation)
            """)
    
            # 5) 🚨 Stoppage Alerts (Improved Table)
            st.markdown("### 🚨 Stoppage Alert Reporting")
            
            if "results" in st.session_state:
                results = st.session_state.results
                df_vis = results.get("df", pd.DataFrame()).copy()
            
                # --- Read threshold values from sidebar ---
                threshold_mode = st.session_state.get("threshold_mode")
                threshold = st.session_state.get("threshold")
                threshold_label = st.session_state.get("threshold_label")
            
                if threshold is None:
                    st.warning("⚠️ Please set a stoppage threshold in the sidebar.")
                else:
                    # --- Filter stoppages ---
                    if "STOP_EVENT" in df_vis.columns and "CT_diff_sec" in df_vis.columns:
                        stoppage_alerts = df_vis[df_vis["CT_diff_sec"] >= threshold].copy()
            
                        if stoppage_alerts.empty:
                            st.info(f"✅ No stoppage alerts found (≥ {threshold_label}).")
                        else:
                            # Add context columns
                            stoppage_alerts["Shots Since Last Stop"] = stoppage_alerts.groupby(
                                stoppage_alerts["STOP_EVENT"].cumsum()
                            ).cumcount()
                            stoppage_alerts["Duration (min)"] = (stoppage_alerts["CT_diff_sec"] / 60).round(1)
                            stoppage_alerts["Reason"] = "to be added"
                            stoppage_alerts["Alert"] = "🔴"
            
                            # Final clean table
                            table = stoppage_alerts[[
                                "SHOT TIME", "Duration (min)", "Shots Since Last Stop", "Reason", "Alert"
                            ]].rename(columns={"SHOT TIME": "Event Time"})
            
                            st.dataframe(table, width="stretch")

    # ---------- Page 2: Raw & Processed Data ----------
    elif page == "📂 Raw & Processed Data":
        st.title("📋 Raw & Processed Cycle Data")

        results = st.session_state.get("results", {})
        if not results:
            st.info("👈 Please generate a report first from the Analysis Dashboard.")
        else:
            df_res = results.get("df", pd.DataFrame()).copy()
            df_vis = results.get("df", pd.DataFrame()).copy()
            stop_events = results.get("stop_events", 0)

            # --- Summary ---
            st.markdown("### Shot Counts & Efficiency")
            st.table(pd.DataFrame({
                "Total Shot Count": [results.get("total_shots", 0)],
                "Normal Shot Count": [results.get("normal_shots", 0)],
                "Efficiency": [f"{results.get('efficiency', 0)*100:.2f}%"],
                "Stop Count": [stop_events]
            }))


            # --- Production & Downtime Summary ---
            st.markdown("### Production & Downtime Summary")
            st.table(pd.DataFrame({
                "Mode CT": [f"{results.get('mode_ct', 0):.2f}"],
                "Lower Limit": [f"{results.get('lower_limit', 0):.2f}"],
                "Upper Limit": [f"{results.get('upper_limit', 0):.2f}"],
                "Production Time (hrs)": [
                    f"{results.get('production_time', 0)/60:.1f} hrs "
                    f"({results.get('production_time', 0)/results.get('total_runtime', 1)*100:.2f}%)"
                ],
                "Downtime (hrs)": [
                    f"{results.get('downtime', 0)/60:.1f} hrs "
                    f"({results.get('downtime', 0)/results.get('total_runtime', 1)*100:.2f}%)"
                ],
                "Total Run Time (hrs)": [f"{results.get('run_hours', 0):.2f}"],
                "Total Stops": [stop_events]
            }))

            st.markdown("---")

            # --- Supplier / Equipment / Approved CT ---
            df_vis["Supplier Name"] = df_vis.get("SUPPLIER NAME", "not provided")
            df_vis["Equipment Code"] = df_vis.get("EQUIPMENT CODE", "not provided")
            df_vis["Approved CT"] = df_vis.get("APPROVED CT", "not provided")
            
            # --- Enrich cycle data ---
            df_vis["Actual CT"] = df_vis["ACTUAL CT"].round(1)
            df_vis["Time Diff Sec"] = df_vis["CT_diff_sec"].round(2)
            
            # --- Reapply dynamic threshold ---
            threshold = st.session_state.get("threshold", results["upper_limit"])
            threshold_mode = st.session_state.get("threshold_mode", "Multiple of Mode CT")
            
            # If using threshold: mark stoppages ≥ threshold
            if threshold_mode in ["Multiple of Mode CT", "Manual (seconds)"]:
                df_vis["Stop_All"] = np.where(df_vis["Time Diff Sec"] >= threshold, 1, 0)
            else:
                # fallback to tolerance band
                df_vis["Stop_All"] = np.where(
                    (df_vis["Time Diff Sec"] < results["lower_limit"]) | 
                    (df_vis["Time Diff Sec"] > results["upper_limit"]), 1, 0
                )
            
            # Preserve "event vs grey" distinction
            df_vis["Stop_Event"] = np.where(df_vis["Stop_All"] & (df_vis["Stop_All"].shift(fill_value=0) == 0), 1, 0)
            
            # Initialise
            df_vis["Cumulative Count"] = 0.0
            df_vis["Run Duration"] = 0.0
            
            current_sum = 0.0
            for i, row in df_vis.iterrows():
                if row["Stop_Event"] == 1:  # first stop in cluster (red tick)
                    # write run duration into this stop row
                    df_vis.at[i, "Run Duration"] = round(current_sum / 60, 2)
                    current_sum = 0.0  # reset after stop
                    df_vis.at[i, "Cumulative Count"] = 0.0
                elif row["Stop_All"] == 1:  # grey stop (bad shot but not event)
                    df_vis.at[i, "Cumulative Count"] = 0.0
                    df_vis.at[i, "Run Duration"] = 0.0
                else:
                    # accumulate production time
                    current_sum += row["Time Diff Sec"] if pd.notna(row["Time Diff Sec"]) else 0
                    df_vis.at[i, "Cumulative Count"] = round(current_sum / 60, 2)

            # --- Final cleaned table ---
            df_clean = df_vis[[
                "Supplier Name", "Equipment Code", "SHOT TIME",
                "Approved CT", "Actual CT", "Time Diff Sec",
                "Stop_All", "Stop_Event", "Stop_Flag",
                "Cumulative Count", "Run Duration"
            ]].rename(columns={"SHOT TIME": "Shot Time"})

            # --- Interactive data editor ---
            st.markdown("### Cycle Data Table (Processed)")
            st.data_editor(
                df_clean,
                width="stretch",
                column_config={
                    "Stop": st.column_config.CheckboxColumn(
                        "Stop",
                        help="Marked as stoppage event",
                        default=False
                    )
                }
            )

            # --- Download options ---
            # 1) CSV Export
            csv = df_clean.to_csv(index=False).encode("utf-8")
            st.download_button(
                label="💾 Download Processed Data (CSV)",
                data=csv,
                file_name="processed_cycle_data.csv",
                mime="text/csv"
            )

            # 2) Excel Export with formulas
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import PatternFill
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            
            def export_to_excel(df, results):
                wb = Workbook()
            
                # ---------------- Sheet 1: Dashboard ----------------
                ws_dash = wb.active
                ws_dash.title = "Dashboard"
            
                ws_dash.append(["📊 Shot Counts & Efficiency"])
                ws_dash.append(["Total Shot Count", results.get("total_shots", 0)])
                ws_dash.append(["Normal Shot Count", results.get("normal_shots", 0)])
                ws_dash.append(["Bad Shot Count", results.get("bad_shots", 0)])
                ws_dash.append(["Efficiency (%)",
                                round((results.get("normal_shots", 0) / results.get("total_shots", 1)) * 100, 2)])
                ws_dash.append(["Stop Count", results.get("stop_events", 0)])
                ws_dash.append([])
            
                ws_dash.append(["⏱ Production & Downtime Summary"])
                ws_dash.append(["Mode CT (sec)", round(results.get("mode_ct", 0), 2)])
                ws_dash.append(["Lower Limit (sec)", round(results.get("lower_limit", 0), 2)])
                ws_dash.append(["Upper Limit (sec)", round(results.get("upper_limit", 0), 2)])
                ws_dash.append(["Production Time (hrs)",
                                f"{results.get('production_time', 0)/60:.2f} hrs "
                                f"({results.get('production_time', 0)/results.get('total_runtime', 1)*100:.2f}%)"])
                ws_dash.append(["Downtime (hrs)",
                                f"{results.get('downtime', 0)/60:.2f} hrs "
                                f"({results.get('downtime', 0)/results.get('total_runtime', 1)*100:.2f}%)"])
                ws_dash.append(["Total Run Time (hrs)", f"{results.get('run_hours', 0):.2f}"])
                ws_dash.append(["Total Stops", results.get("stop_events", 0)])
            
                for col in ws_dash.columns:
                    max_len = max(len(str(c.value)) if c.value else 0 for c in col)
                    ws_dash.column_dimensions[col[0].column_letter].width = max_len + 2
            
                # ---------------- Sheet 2: Processed Data ----------------
                ws_data = wb.create_sheet("Processed Data")
            
                # keep only existing columns in the specified order
                cols_to_keep = [
                    "Shot Time", "Supplier Name", "Equipment Code", "Approved CT",
                    "Actual CT", "Time Diff Sec", "Stop_All", "Stop_Event",
                    "Stop_Flag", "Cumulative Count", "Run Duration"
                ]
                existing_cols = [c for c in cols_to_keep if c in df.columns]
                df_export = df[existing_cols]
            
                # header
                ws_data.append(list(df_export.columns))
            
                # precompute column letters by name (works for any order)
                def col_letter(col_name):
                    return get_column_letter(df_export.columns.get_loc(col_name) + 1)
            
                # these names must exist for formulas
                td_col = col_letter("Time Diff Sec")               # seconds
                se_col = col_letter("Stop_Event")                  # 1/0
                cc_col = col_letter("Cumulative Count")            # minutes (computed)
                rd_col = col_letter("Run Duration")                # minutes (computed)
            
                # write rows with formulas for cumulative & run duration
                for r_idx, row_vals in enumerate(dataframe_to_rows(df_export, index=False, header=False), start=2):
                    row_out = []
                    for c_idx, value in enumerate(row_vals, 1):
                        header = df_export.columns[c_idx - 1]
            
                        if header == "Cumulative Count":
                            if r_idx == 2:
                                # first data row: no previous cumulative
                                row_out.append(f"=IF({se_col}{r_idx}=1,0,IF({td_col}{r_idx}=\"\",0,{td_col}{r_idx}/60))")
                            else:
                                row_out.append(
                                    f"=IF({se_col}{r_idx}=1,0,"
                                    f"IF({cc_col}{r_idx-1}=\"\",0,{cc_col}{r_idx-1})+IF({td_col}{r_idx}=\"\",0,{td_col}{r_idx}/60))"
                                )
            
                        elif header == "Run Duration":
                            # at stop event row, show the run duration accumulated BEFORE this row
                            if r_idx == 2:
                                row_out.append(f"=IF({se_col}{r_idx}=1,0,0)")
                            else:
                                row_out.append(f"=IF({se_col}{r_idx}=1,IF({cc_col}{r_idx-1}=\"\",0,{cc_col}{r_idx-1}),0)")
            
                        else:
                            row_out.append(value)
            
                    ws_data.append(row_out)
            
                ws_data.freeze_panes = "A2"
            
                # number format for the two computed columns
                cc_idx = df_export.columns.get_loc("Cumulative Count") + 1 if "Cumulative Count" in df_export.columns else None
                rd_idx = df_export.columns.get_loc("Run Duration") + 1 if "Run Duration" in df_export.columns else None
                if cc_idx:
                    for r in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, min_col=cc_idx, max_col=cc_idx):
                        r[0].number_format = "0.00"
                if rd_idx:
                    for r in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, min_col=rd_idx, max_col=rd_idx):
                        r[0].number_format = "0.00"
            
                # Highlight stops
                grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
                if "Stop_All" in df_export.columns:
                    idx = df_export.columns.get_loc("Stop_All") + 1
                    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row):
                        if row[idx - 1].value == 1:
                            row[idx - 1].fill = grey_fill
            
                if "Stop_Event" in df_export.columns:
                    idx = df_export.columns.get_loc("Stop_Event") + 1
                    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row):
                        if row[idx - 1].value == 1:
                            row[idx - 1].fill = red_fill
            
                # autosize
                for col in ws_data.columns:
                    max_len = max(len(str(c.value)) if c.value else 0 for c in col)
                    ws_data.column_dimensions[col[0].column_letter].width = max_len + 2
            
                # save
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                return buffer
            
            # --- Inside your app ---
            excel_buffer = export_to_excel(df_vis, results)
            st.download_button(
                label="📊 Download Excel Report (with Dashboard)",
                data=excel_buffer,
                file_name="processed_cycle_data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # --- Page 3: Daily Analysis (selected week) ---
    elif page == "📅 Daily Analysis":
        st.title("📅 Daily Trends Dashboard")
        st.subheader(f"Tool: {tool}")
    
        results = st.session_state.get("results", {})
        if not results or "df" not in results:
            st.info("👈 Please generate a report first from the sidebar.")
            st.stop()
    
        df = results["df"].copy()
        df["DAY"] = df["SHOT TIME"].dt.date
    
        # --- Select Week ---
        min_date, max_date = df["SHOT TIME"].dt.date.min(), df["SHOT TIME"].dt.date.max()
        week_options = pd.date_range(min_date, max_date, freq="W-MON").date
        selected_week = st.selectbox(
            "Select Week",
            week_options,
            format_func=lambda d: f"Week of {d}"
        )
    
        week_mask = (df["DAY"] >= selected_week) & (df["DAY"] < (selected_week + timedelta(days=7)))
        df_week = df.loc[week_mask].copy()
    
        if df_week.empty:
            st.warning("No data for this week.")
            st.stop()
    
        # --- Daily Summary Table (Page 1 logic reused per day) ---
        def calc_metrics(g):
            stops = g["STOP_EVENT"].sum()
            if stops > 0:
                downtime_events = g.loc[g["STOP_EVENT"], "CT_diff_sec"] / 60
                mttr = downtime_events.mean() if not downtime_events.empty else np.nan
    
                total_uptime = g["CT_diff_sec"].sum() / 60
                mtbf = total_uptime / stops
            else:
                mttr, mtbf = np.nan, np.nan
    
            return pd.Series({
                "Total Shots": len(g),
                "Normal Shots": ((g["CT_diff_sec"] >= results["lower_limit"]) &
                                 (g["CT_diff_sec"] <= results["upper_limit"])).sum(),
                "Bad Shots": ((g["CT_diff_sec"] < results["lower_limit"]) |
                              (g["CT_diff_sec"] > results["upper_limit"])).sum(),
                "Stops": stops,
                "MTTR (min)": mttr,
                "MTBF (min)": mtbf
            })
    
        daily_summary = df_week.groupby("DAY").apply(calc_metrics).reset_index()
    
        # Efficiency
        daily_summary["Efficiency (%)"] = (
            daily_summary["Normal Shots"] / daily_summary["Total Shots"] * 100
        ).round(2)
    
        # Stability Index
        daily_summary["Stability Index (%)"] = (
            daily_summary.apply(
                lambda r: (r["MTBF (min)"] / (r["MTBF (min)"] + r["MTTR (min)"])) * 100
                if pd.notna(r["MTBF (min)"]) and pd.notna(r["MTTR (min)"]) else np.nan,
                axis=1
            )
        ).round(2)
    
        # Force SI = 100% if no stops
        daily_summary.loc[daily_summary["Stops"] == 0, "Stability Index (%)"] = 100.0
    
        st.markdown("### 📋 Weekly Summary Table (Daily Breakdown)")
        st.dataframe(daily_summary, use_container_width=True)
    
        # --- Reuse Page-1 style charts but grouped by DAY ---
    
        # Use run end timestamps to keep only runs that ended inside the selected week
        run_durations_all = results["run_durations"].copy()
        bucket_order = results["bucket_order"]
        bucket_color_map = results["bucket_color_map"]
    
        rd_week = run_durations_all.loc[
            (run_durations_all["RUN_END"].dt.date >= selected_week) &
            (run_durations_all["RUN_END"].dt.date <= selected_week + timedelta(days=6))
        ].copy()
    
        # 1) Time Bucket Analysis
        bucket_counts = (
            rd_week["TIME_BUCKET"]
            .value_counts()
            .reindex(bucket_order)
            .fillna(0)
            .astype(int)
        )
        total_runs = int(bucket_counts.sum()) if not bucket_counts.empty else 0
        bucket_df = bucket_counts.reset_index()
        bucket_df.columns = ["Time Bucket", "Occurrences"]
        bucket_df["Percentage"] = np.where(
            total_runs > 0, (bucket_df["Occurrences"] / total_runs * 100).round(2), 0.0
        )
    
        fig_bucket = px.bar(
            bucket_df[bucket_df["Time Bucket"].notna()],
            x="Occurrences", y="Time Bucket", orientation="h", text="Occurrences",
            title="Time Bucket Analysis (Selected Week)",
            category_orders={"Time Bucket": bucket_order},
            color="Time Bucket",
            color_discrete_map=bucket_color_map,
            hover_data={"Occurrences": True, "Percentage": True}
        )
        fig_bucket.update_traces(textposition="outside")
        fig_bucket.update_layout(legend=dict(orientation="h", x=0.5, y=-0.25, xanchor="center"))
        st.plotly_chart(fig_bucket, use_container_width=True)
    
        with st.expander("📊 Time Bucket Analysis – Data", expanded=False):
            st.dataframe(bucket_df, use_container_width=True)
    
        # 2) Daily Time Bucket Trend
        rd_week["DAY"] = rd_week["RUN_END"].dt.date
        trend = rd_week.groupby(["DAY", "TIME_BUCKET"]).size().reset_index(name="count")
    
        days = pd.date_range(selected_week, selected_week + timedelta(days=6)).date
        grid = pd.MultiIndex.from_product([days, bucket_order], names=["DAY","TIME_BUCKET"]).to_frame(index=False)
        trend = grid.merge(trend, on=["DAY","TIME_BUCKET"], how="left").fillna({"count": 0})
    
        fig_tb_trend = px.bar(
            trend, x="DAY", y="count", color="TIME_BUCKET",
            category_orders={"TIME_BUCKET": bucket_order},
            color_discrete_map=bucket_color_map,
            title="Daily Time Bucket Trend (Selected Week)",
            hover_data={"count": True, "DAY": True}
        )
        fig_tb_trend.update_layout(
            barmode="stack",
            legend=dict(orientation="h", x=0.5, y=-0.25, xanchor="center"),
            xaxis=dict(title="Day")
        )
        st.plotly_chart(fig_tb_trend, use_container_width=True)
    
        with st.expander("📊 Daily Time Bucket Trend – Data", expanded=False):
            st.dataframe(trend, use_container_width=True)
    
        # 3) MTTR & MTBF Trend by Day
        fig_mt = go.Figure()
        x_days = daily_summary["DAY"]
    
        fig_mt.add_trace(go.Scatter(
            x=x_days, y=daily_summary["MTTR (min)"],
            mode="lines+markers", name="MTTR (min)",
            line=dict(color="red", width=2), yaxis="y"
        ))
    
        fig_mt.add_trace(go.Scatter(
            x=x_days, y=daily_summary["MTBF (min)"],
            mode="lines+markers", name="MTBF (min)",
            line=dict(color="green", width=2, dash="dot"), yaxis="y2"
        ))
    
        fig_mt.update_layout(
            title="MTTR & MTBF Trend by Day",
            xaxis=dict(title="Day"),
            yaxis=dict(title="MTTR (min)", tickfont=dict(color="red"), side="left"),
            yaxis2=dict(title="MTBF (min)", tickfont=dict(color="green"), overlaying="y", side="right"),
            margin=dict(l=60, r=60, t=60, b=40),
            legend=dict(orientation="h", x=0.5, y=-0.25, xanchor="center"),
            showlegend=True
        )
        st.plotly_chart(fig_mt, use_container_width=True)
    
        with st.expander("📊 MTTR & MTBF – Data", expanded=False):
            st.dataframe(
                daily_summary[["DAY","MTTR (min)","MTBF (min)","Stops"]].copy(),
                use_container_width=True
            )
    
        # 4) Stability Index Trend by Day
        fig_stability = go.Figure()
        fig_stability.add_trace(go.Scatter(
            x=daily_summary["DAY"].astype(str),
            y=daily_summary["Stability Index (%)"],
            mode="lines+markers",
            name="Stability Index (%)",
            line=dict(color="blue", width=2)
        ))
    
        for y0, y1, c in [(0,50,"red"),(50,70,"yellow"),(70,100,"green")]:
            fig_stability.add_shape(
                type="rect",
                xref="paper", x0=0, x1=1,
                y0=y0, y1=y1,
                fillcolor=c, opacity=0.1, line_width=0, layer="below"
            )
    
        fig_stability.update_layout(
            title="Stability Index by Day",
            xaxis=dict(title="Day", tickmode="array", tickvals=list(range(len(daily_summary))),
                       ticktext=daily_summary["DAY"].astype(str).tolist()),
            yaxis=dict(title="Stability Index (%)", range=[0,100]),
            margin=dict(l=60,r=60,t=60,b=40),
            legend=dict(orientation="h", x=0.5, y=-0.25, xanchor="center")
        )
        st.plotly_chart(fig_stability, use_container_width=True)
    
        with st.expander("📊 Stability Index – Data", expanded=False):
            st.dataframe(
                daily_summary[["DAY","Stability Index (%)","MTTR (min)","MTBF (min)","Stops"]].copy(),
                use_container_width=True
            )
    
    
else:
    st.info("👈 Upload a cleaned run rate Excel file to begin. Headers in ROW 1 please.")